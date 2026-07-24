"""
=========================================================
Attendance Notification System Pro
Report Generator
Version : 10.0 Enterprise
=========================================================
"""

import os

import pandas as pd

from openpyxl import load_workbook

from openpyxl.styles import (

    Font,

    PatternFill,

    Alignment,

    Border,

    Side

)

from config import REPORT_FOLDER


class ReportGenerator:
    """
    Enterprise Excel Report Generator

    Features
    --------
    • Professional Excel Report
    • Auto Column Width
    • Status Highlight
    • Freeze Header
    • Auto Filter
    • Border Formatting
    • High Performance
    """

    # =====================================================
    # Initialize
    # =====================================================

    def __init__(self):

        os.makedirs(

            REPORT_FOLDER,

            exist_ok=True

        )

        self.header_fill = PatternFill(

            fill_type="solid",

            start_color="1F4E78",

            end_color="1F4E78"

        )

        self.header_font = Font(

            bold=True,

            color="FFFFFF"

        )

        self.center = Alignment(

            horizontal="center",

            vertical="center"

        )

        self.green_fill = PatternFill(

            fill_type="solid",

            start_color="C6EFCE",

            end_color="C6EFCE"

        )

        self.yellow_fill = PatternFill(

            fill_type="solid",

            start_color="FFF2CC",

            end_color="FFF2CC"

        )

        self.orange_fill = PatternFill(

            fill_type="solid",

            start_color="FCE4D6",

            end_color="FCE4D6"

        )

        self.red_fill = PatternFill(

            fill_type="solid",

            start_color="F4CCCC",

            end_color="F4CCCC"

        )

        thin = Side(

            border_style="thin",

            color="BFBFBF"

        )

        self.border = Border(

            left=thin,

            right=thin,

            top=thin,

            bottom=thin

        )
        # =====================================================
    # Generate Excel Report
    # =====================================================

    def generate_excel(
        self,
        employees,
        filename
    ):

        if not employees:

            raise ValueError(
                "No employee data available."
            )

        report_path = os.path.join(

            REPORT_FOLDER,

            filename

        )

        print("=" * 60)
        print("Generating Excel Report...")
        print("=" * 60)

        # -----------------------------------------
        # Employee Data
        # -----------------------------------------

        dataframe = pd.DataFrame([

            {

                "Employee ID": emp.get(

                    "employee_id",

                    ""

                ),

                "Employee Name": emp.get(

                    "name",

                    ""

                ),

                "Department": emp.get(

                    "department",

                    ""

                ),

                "Designation": emp.get(

                    "designation",

                    ""

                ),

                "Attendance Date": emp.get(

                    "attendance_date",

                    ""

                ),

                "Punch In": emp.get(

                    "punch_in",

                    "--"

                ),

                "Punch Out": emp.get(

                    "punch_out",

                    "--"

                ),

                "Daily OT": emp.get(

                    "daily_ot",

                    "00:00"

                ),

                "Monthly OT": emp.get(

                    "monthly_ot",

                    "00:00"

                ),

                "Remaining OT": emp.get(

                    "remaining_ot",

                    "00:00"

                ),

                "Daily Status": emp.get(

                    "daily_status",

                    "Normal"

                ),

                "Monthly Status": emp.get(

                    "monthly_status",

                    "Normal"

                )

            }

            for emp in employees

        ])

        dataframe.to_excel(

            report_path,

            index=False,

            engine="openpyxl"

        )

        workbook = load_workbook(

            report_path

        )
        worksheet = workbook.active
        worksheet.title = "Attendance Report"
        # =====================================================
        # Header Formatting
        # =====================================================

        for cell in worksheet[1]:

            cell.fill = self.header_fill

            cell.font = self.header_font

            cell.alignment = self.center

            cell.border = self.border

        # =====================================================
        # Freeze Header
        # =====================================================

        worksheet.freeze_panes = "A2"

        # =====================================================
        # Auto Filter
        # =====================================================

        worksheet.auto_filter.ref = worksheet.dimensions

        # =====================================================
        # Row Formatting
        # =====================================================

        for row in worksheet.iter_rows(

            min_row=2

        ):

            for cell in row:

                cell.alignment = self.center

                cell.border = self.border

        # =====================================================
        # Row Height
        # =====================================================

        worksheet.row_dimensions[1].height = 28

        for row in range(

            2,

            worksheet.max_row + 1

        ):

            worksheet.row_dimensions[row].height = 22
            # =====================================================
        # Auto Column Width
        # =====================================================

        for column_cells in worksheet.columns:

            column_letter = column_cells[0].column_letter

            max_length = len(str(column_cells[0].value))

            for cell in column_cells:

                try:

                    if cell.value is not None:

                        value_length = len(str(cell.value))

                        if value_length > max_length:

                            max_length = value_length

                except Exception:

                    pass

            worksheet.column_dimensions[

                column_letter

            ].width = min(

                max_length + 4,

                35

            )

        # =====================================================
        # Status Cell Coloring
        # =====================================================

        daily_status_column = None

        monthly_status_column = None

        for cell in worksheet[1]:

            if cell.value == "Daily Status":

                daily_status_column = cell.column

            elif cell.value == "Monthly Status":

                monthly_status_column = cell.column

        for row in range(

            2,

            worksheet.max_row + 1

        ):

            # -----------------------------------------
            # Daily Status
            # -----------------------------------------

            if daily_status_column:

                cell = worksheet.cell(

                    row=row,

                    column=daily_status_column

                )

                value = str(cell.value).strip()

                if value == "Normal":

                    cell.fill = self.green_fill

                elif value == "Warning":

                    cell.fill = self.yellow_fill

                elif value == "Limit Reached":

                    cell.fill = self.orange_fill

                elif value == "Exceeded":

                    cell.fill = self.red_fill

                cell.alignment = self.center

                cell.border = self.border

            # -----------------------------------------
            # Monthly Status
            # -----------------------------------------

            if monthly_status_column:

                cell = worksheet.cell(

                    row=row,

                    column=monthly_status_column

                )

                value = str(cell.value).strip()

                if value == "Normal":

                    cell.fill = self.green_fill

                elif value == "Warning":

                    cell.fill = self.yellow_fill

                elif value == "Limit Reached":

                    cell.fill = self.orange_fill

                elif value == "Exceeded":

                    cell.fill = self.red_fill

                cell.alignment = self.center

                cell.border = self.border
        
        # =====================================================
        # Create Summary Sheet
        # =====================================================

        summary_sheet = workbook.create_sheet(

            title="Summary"

        )
        summary_sheet = workbook["Summary"]

        total = len(employees)

        present = sum(
            1
            for emp in employees
            if not any(
                keyword in str(status)
                for keyword in (
                    "Late Punch",
                    "Early Out",
                    "Missing Punch"
                )
                for status in emp.get("status", [])
            )
        )
        late = sum(

            1

            for emp in employees

            if any(

                "Late Punch" in str(status)

                for status in emp.get(

                    "status",

                    []

                )

            )

        )

        early = sum(

            1

            for emp in employees

            if any(

                "Early Out" in str(status)

                for status in emp.get(

                    "status",

                    []

                )

            )

        )

        overtime = sum(

            1

            for emp in employees

            if emp.get(

                "daily_ot_minutes",

                0

            ) > 0

        )

        warning = sum(

            1

            for emp in employees

            if emp.get(

                "monthly_status"

            ) == "Warning"

        )

        limit = sum(

            1

            for emp in employees

            if emp.get(

                "monthly_status"

            ) == "Limit Reached"

        )

        exceeded = sum(

            1

            for emp in employees

            if emp.get(

                "monthly_status"

            ) == "Exceeded"

        )

        summary_sheet.append(

            [

                "Attendance Notification System Pro"

            ]

        )
        summary_sheet["A1"].font = Font(
            bold=True,
            size=16,
            color="1F4E78"
        )
        
        summary_sheet["A1"].alignment = self.center
        
        summary_sheet.merge_cells("A1:B1")

        summary_sheet.append([])

        summary_sheet.append(

            [

                "Attendance Summary",

                "Count"

            ]

        )

        summary_sheet.append(

            [

                "Total Employees",

                total

            ]

        )

        summary_sheet.append(

            [

                "Present",

                present

            ]

        )

        summary_sheet.append(

            [

                "Late Punch",

                late

            ]

        )

        summary_sheet.append(

            [

                "Early Out",

                early

            ]

        )

        summary_sheet.append(

            [

                "Overtime Employees",

                overtime

            ]

        )

        summary_sheet.append(

            [

                "OT Warning",

                warning

            ]

        )

        summary_sheet.append(

            [

                "Limit Reached",

                limit

            ]

        )

        summary_sheet.append(

            [

                "OT Exceeded",

                exceeded

            ]

        )

        # -----------------------------------------
        # Summary Sheet Formatting
        # -----------------------------------------

        for cell in summary_sheet[3]:

            cell.fill = self.header_fill

            cell.font = self.header_font

            cell.alignment = self.center

            cell.border = self.border

        for row in summary_sheet.iter_rows(

            min_row=4

        ):

            for cell in row:

                cell.alignment = self.center

                cell.border = self.border
                
            if row[0].value == "Total Employees":
                row[1].fill = self.header_fill
                row[1].font = self.header_font

            elif row[0].value == "Present":

                row[1].fill = self.green_fill

            elif row[0].value == "OT Warning":

                row[1].fill = self.yellow_fill

            elif row[0].value == "Limit Reached":

                row[1].fill = self.orange_fill

            elif row[0].value == "OT Exceeded":

                row[1].fill = self.red_fill

        summary_sheet.column_dimensions["A"].width = 35

        summary_sheet.column_dimensions["B"].width = 18
        
        summary_sheet.auto_filter.ref = summary_sheet.dimensions
        
        summary_sheet.freeze_panes = "A4"
        
        
        # =====================================================
        # Activate Attendance Sheet
        # =====================================================

        summary_index = workbook.sheetnames.index("Summary")

        workbook.active = summary_index

        # =====================================================
        # Save Workbook
        # =====================================================

        workbook.save(

            report_path

        )

        workbook.close()

        print("=" * 60)
        print("Excel Report Generated Successfully")
        print("=" * 60)
        print(f"Report Path : {report_path}")
        print("=" * 60)

        return report_path